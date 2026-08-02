"""
Extracts clean monthly revenue targets from all 13 rep sheets in the
original workbook. Bypasses the fragile Summary-tab aggregation formulas
(which contain #REF! errors) by reading each rep sheet's own
'Total Month Revenue' rows directly, which resolve correctly.
"""
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

SOURCE = "/mnt/user-data/uploads/PRODUCT_SALES_TARGETS_2024_-_2025_Rev3_20250811.xlsm"

MONTH_ABBR_FIX = {
    "sept": "sep",
}

def parse_month_cell(value):
    """Best-effort parse of a single month-header cell into a normalized
    first-of-month datetime. Source sheets mix real datetime cells with
    inconsistently formatted text labels (e.g. 'Sept-25', 'Apr 2026',
    'Jul 2025') -- this handles all of them. Returns None if unparseable."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)
    if isinstance(value, str):
        s = value.strip().lower().replace(",", "")
        for wrong, right in MONTH_ABBR_FIX.items():
            s = s.replace(wrong, right)
        # try common explicit patterns first: "sep-25", "sep 2025", "sep2025"
        m = re.match(r"([a-z]{3,9})[\s\-]?(\d{2,4})", s)
        if m:
            mon_str, year_str = m.groups()
            try:
                mon_num = datetime.strptime(mon_str[:3], "%b").month
                year = int(year_str)
                if year < 100:
                    year += 2000
                return datetime(year, mon_num, 1)
            except ValueError:
                pass
        try:
            import dateutil.parser
            dt = dateutil.parser.parse(value, fuzzy=True, default=datetime(1900, 1, 1))
            return datetime(dt.year, dt.month, 1)
        except Exception:
            return None
    return None

def resolve_block_months(raw_values):
    """Given the 12 raw header-row cell values for one block, return a
    clean list of 12 normalized month dates -- filling any unparseable
    cells by extrapolating from the ones that did parse, since each block
    is always 12 consecutive calendar months."""
    parsed = [parse_month_cell(v) for v in raw_values]
    # find an anchor: first successfully parsed cell and its position
    anchor_idx, anchor_month = None, None
    for i, p in enumerate(parsed):
        if p is not None:
            anchor_idx, anchor_month = i, p
            break
    if anchor_month is None:
        return [None] * 12  # entire block unparseable -- skip it
    resolved = []
    for i in range(12):
        offset = i - anchor_idx
        year = anchor_month.year + (anchor_month.month - 1 + offset) // 12
        month = (anchor_month.month - 1 + offset) % 12 + 1
        resolved.append(datetime(year, month, 1))
    return resolved



# Rep sheet tab name -> clean display name (normalizes mismatches with actuals export)
REP_SHEETS = {
    "AnaR": "Ana Roque",
    "JHB2": "Joburg Account Manager 2",
    "LeanetteM": "Leanette Mtsweni",
    "AndreaK": "Andrea Klopper",
    "NtokozoM": "Ntokozo Masango",
    "ShetiR": "Sheti Ramokone",
    "KishaE": "Kisha Edwards",
    "ThatoM": "Thato Moratho",
    "AM - 8": "AM 8",
    "AM - 9": "AM 9",
    "AM 10": "AM 10",
    "CSA-Telesales 1": "CSA / Telesales 1",
    "CSA-Telesales 2": "CSA / Telesales 2",
    "CSA-Telesales 3": "CSA / Telesales 3",
    "CSA-Telesales 4": "CSA / Telesales 4",
    "HouseAccount": "House Account",
}

# Row where each category block's "Applicable" / category label sits,
# and the offset (+N rows) to that category's "Total Month Revenue" line.
# Identified by scanning column C for category headers and "Total Month Revenue".
CATEGORY_HEADER_ROW_HINTS = [
    "Records Management - New",
    "Records Management - Ongoing",
    "Backup Storage & Management",
    "Image Processing",
    "Software Integration and Related Services",
]

def find_category_blocks(ws):
    """
    Scan column C top to bottom and pair each category header row with the
    NEXT 'Total Month Revenue' row that follows it (in document order).
    Each rep sheet contains TWO stacked blocks per category (one per
    target year), so this returns a list of (category, header_row, total_row)
    tuples rather than a dict, preserving both occurrences instead of the
    second silently overwriting the first.
    """
    blocks = []
    current_category = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=250, max_col=3):
        c_cell = row[2]  # column C
        val = c_cell.value
        if val in CATEGORY_HEADER_ROW_HINTS:
            current_category = val
            header_row = c_cell.row
        elif val == "Total Month Revenue" and current_category:
            blocks.append((current_category, header_row, c_cell.row))
            current_category = None
    return blocks

def extract_rep_targets(sheet_name, display_name, wb):
    ws = wb[sheet_name]
    blocks = find_category_blocks(ws)
    records = []
    month_col_start = 5  # columns E through P, 12 months per block
    for category, header_row, total_row in blocks:
        raw_month_cells = [ws.cell(row=header_row, column=month_col_start + i).value for i in range(12)]
        resolved_months = resolve_block_months(raw_month_cells)
        for i in range(12):
            month_date = resolved_months[i]
            if month_date is None:
                continue
            col = month_col_start + i
            target_value = ws.cell(row=total_row, column=col).value
            records.append({
                "Rep": display_name,
                "Sheet": sheet_name,
                "Category": category,
                "Month": month_date,
                "Target Revenue (ZAR)": target_value if isinstance(target_value, (int, float)) else 0,
            })
    return records

def main():
    wb = load_workbook(SOURCE, data_only=True)
    all_records = []
    for sheet_name, display_name in REP_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found, skipping")
            continue
        recs = extract_rep_targets(sheet_name, display_name, wb)
        all_records.extend(recs)
        print(f"{sheet_name:20s} -> {display_name:30s}  {len(recs)} target rows extracted")

    df = pd.DataFrame(all_records)
    out_path = "/home/claude/sales_pipeline/clean_targets.csv"
    df.to_csv(out_path, index=False)
    print(f"\nSaved {len(df)} rows to {out_path}")
    print(df.head(20))
    return df

if __name__ == "__main__":
    main()
