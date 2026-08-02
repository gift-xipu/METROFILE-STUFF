"""
Send the monthly rate revision pack via Outlook - THE ACTUAL "BUTTON".
==========================================================================
Run this AFTER generate_monthly_pack.py, on a Windows machine that has
Microsoft Outlook installed and configured with the sender's mailbox.

It reads output/email_manifest.xlsx and, for every row, creates and sends
an Outlook email to the sales person with the letter + rate schedule
attached. It drives your real, already-logged-in Outlook desktop app - no
passwords, API keys or admin setup required.

SETUP (one-time):
    pip install pywin32

USAGE:
    python send_via_outlook.py                 # sends every row in the manifest
    python send_via_outlook.py --dry-run        # prints what would be sent, sends nothing
    python send_via_outlook.py --draft-only      # creates the emails in Outlook as
                                                    Drafts instead of sending, so a
                                                    manager can review/approve first

TO TURN THIS INTO A ONE-CLICK BUTTON:
    - Windows: save a .bat file next to this script containing:
          python generate_monthly_pack.py --month "%1"
          python send_via_outlook.py --draft-only
      then either double-click it or trigger it from Task Scheduler on the
      1st of every month with the right target month.
    - Excel/Outlook: wrap the same two calls in a VBA Sub behind a button on
      the tracker workbook using Shell(), so the whole thing runs from a
      literal button inside the spreadsheet sales already looks at.
    - Fully server-side (no one's PC needs to be on): move the send step to
      Microsoft Graph API (send mail as a shared mailbox) triggered by Power
      Automate on a schedule - the natural next step once this proves out.
"""
import argparse
import sys

import openpyxl

MANIFEST_PATH = "output/email_manifest.xlsx"


def load_manifest():
    wb = openpyxl.load_workbook(MANIFEST_PATH, data_only=True)
    ws = wb.active
    header_row = 2
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def send_all(dry_run=False, draft_only=False):
    """Send/draft/preview every row in the manifest. Returns a list of
    {client, sales_person, email, status} dicts describing what happened."""
    rows = load_manifest()
    results = []

    if dry_run:
        for r in rows:
            results.append({"client": r["Client Name"], "sales_person": r["Sales Person"],
                             "email": r["Sales Person Email"], "status": "Would send (dry run)"})
        return results

    try:
        import win32com.client
    except ImportError:
        raise RuntimeError(
            "pywin32 is not installed, or this isn't running on Windows with Outlook "
            "installed. Run:  pip install pywin32"
        )

    outlook = win32com.client.Dispatch("Outlook.Application")

    for r in rows:
        mail = outlook.CreateItem(0)  # olMailItem
        mail.To = r["Sales Person Email"]
        mail.Subject = r["Subject"]
        mail.Body = r["Body"]
        mail.Attachments.Add(r["Letter Attachment"])
        mail.Attachments.Add(r["Schedule Attachment"])

        if draft_only:
            mail.Save()
            results.append({"client": r["Client Name"], "sales_person": r["Sales Person"],
                             "email": r["Sales Person Email"], "status": "Draft created"})
        else:
            mail.Send()
            results.append({"client": r["Client Name"], "sales_person": r["Sales Person"],
                             "email": r["Sales Person Email"], "status": "Sent"})

    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Preview without sending")
    parser.add_argument("--draft-only", action="store_true", help="Save as Outlook drafts instead of sending")
    args = parser.parse_args()
    for r in send_all(dry_run=args.dry_run, draft_only=args.draft_only):
        print(f"  {r['status']}: {r['email']} ({r['client']})")
