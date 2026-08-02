"""
Builds master_rates.xlsx - the central rate table that the generation pipeline
reads from. One sheet, long format: PriceLevel | Category | Item | ChargeCode | Rate | Period

Seeded with the REAL Standard/Master rates (POL0000012) taken from the uploaded
rate card PDF. Every other price level currently in the tracker is auto-derived
from the standard rates with a small placeholder variance, purely so the pipeline
has something to generate against today. Replace DEMO_VARIANCE rows with your
actual negotiated rates as you capture them from the individual client agreements.
"""
import csv
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl as oxl

random.seed(7)

PERIOD = "Jul 2026 to Jun 2027"

# ---- Load the real standard rate card ----
standard_rows = []
with open("standard_rates_seed.csv") as f:
    reader = csv.DictReader(f)
    for r in reader:
        standard_rows.append({
            "PriceLevel": r["PriceLevel"],
            "Category": r["Category"],
            "Item": r["Item"],
            "ChargeCode": r["ChargeCode"],
            "Rate": float(r["Rate"]),
        })

# ---- Pull the distinct price levels currently in the tracker ----
wb_tracker = openpyxl.load_workbook("Rate_Revision_Schedule.xlsx", data_only=True)
ws_tracker = wb_tracker.active
price_levels = []
seen = set()
for row in ws_tracker.iter_rows(min_row=6, max_row=ws_tracker.max_row, values_only=True):
    pl = row[0]
    if pl and pl not in seen:
        seen.add(pl)
        price_levels.append(pl)

# ---- Build the full long-format rate table ----
all_rows = []
for r in standard_rows:
    all_rows.append({**r, "Period": PERIOD, "Source": "Standard rate card (POL0000012)"})

for pl in price_levels:
    if pl == "POL0000012":
        continue
    variance = round(random.uniform(-0.08, 0.05), 3)  # placeholder negotiated variance
    for r in standard_rows:
        all_rows.append({
            "PriceLevel": pl,
            "Category": r["Category"],
            "Item": r["Item"],
            "ChargeCode": r["ChargeCode"],
            "Rate": round(r["Rate"] * (1 + variance), 2),
            "Period": PERIOD,
            "Source": "DEMO PLACEHOLDER \u2013 replace with actual negotiated rate",
        })

# ---- Write workbook ----
NAVY = "1F3864"; STEEL = "2E5395"; LIGHT_GREY = "F2F2F2"; WHITE = "FFFFFF"
FONT = "Arial"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rates"
ws.sheet_view.showGridLines = False

headers = ["Price Level", "Category", "Item", "Charge Code", "Rate (ZAR)", "Period", "Source"]
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
t = ws.cell(row=1, column=1, value="MASTER RATE TABLE \u2013 All Price Levels")
t.font = Font(name=FONT, size=14, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
s = ws.cell(row=2, column=1, value=("One row per Price Level per Charge Code. This is the single source of truth the "
    "generation script reads rates from. Rows marked 'DEMO PLACEHOLDER' are illustrative only \u2013 "
    "replace their Rate values with the real negotiated figures from each client's agreement."))
s.font = Font(name=FONT, size=9.5, italic=True, color=WHITE)
s.fill = PatternFill("solid", fgColor=STEEL)
s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ws.row_dimensions[2].height = 28

header_row = 3
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[header_row].height = 20

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

data_start = header_row + 1
for i, r in enumerate(all_rows):
    row_num = data_start + i
    band = PatternFill("solid", fgColor=WHITE if i % 2 == 0 else LIGHT_GREY)
    vals = [r["PriceLevel"], r["Category"], r["Item"], r["ChargeCode"], r["Rate"], r["Period"], r["Source"]]
    for col, v in enumerate(vals, start=1):
        cell = ws.cell(row=row_num, column=col, value=v)
        cell.font = Font(name=FONT, size=10, italic=(col == 7))
        cell.fill = band
        cell.border = border
        cell.alignment = Alignment(horizontal="left" if col in (2, 3, 7) else "center", vertical="center")
        if col == 5:
            cell.number_format = '"R" #,##0.00'

widths = {1: 14, 2: 32, 3: 45, 4: 12, 5: 13, 6: 20, 7: 42}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = f"A{data_start}"
ws.auto_filter.ref = f"A{header_row}:G{data_start + len(all_rows) - 1}"

wb.save("master_rates.xlsx")
print(f"master_rates.xlsx built: {len(all_rows)} rate rows across {len(price_levels)} price levels")
