"""
Rate Revision Monthly Pack Generator
=====================================
Run this once a month (or wire it to a scheduler / button) to produce every
letter + rate schedule that needs to go out to sales people that month.

USAGE:
    python3 generate_monthly_pack.py --month "April 2027"

WHAT IT DOES:
    1. Reads Rate_Revision_Schedule.xlsx (the tracker) and finds every client
       whose Notification Due Date falls in the target month.
    2. Reads master_rates.xlsx for that client's Price Level (falls back to the
       Standard rate card POL0000012 if a client has no negotiated rates yet).
    3. Builds a personalised Price Increase Letter (.docx) from letter_template.docx.
    4. Builds a personalised Rate Schedule (.xlsx) styled like the standard rate card.
    5. Saves both into output/<SalesPerson>/<ClientName>/
    6. Writes email_manifest.xlsx - one row per client, with the sales person's
       name/email, subject line, draft body text, and the two attachment paths -
       ready to be picked up by send_via_outlook.py.

WHAT IT DOES NOT DO (yet):
    Actually send the email. That needs to happen from a machine with Outlook
    installed (see send_via_outlook.py) or via the Microsoft Graph API.
"""
import argparse
import os
import shutil
import zipfile
import re
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TRACKER = "Rate_Revision_Schedule.xlsx"
MASTER_RATES = "master_rates.xlsx"
LETTER_TEMPLATE = "letter_template.docx"
OUTPUT_DIR = "output"
STANDARD_LEVEL = "POL0000012"

FONT = "Arial"
NAVY = "1F3864"
STEEL = "2E5395"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"


def safe_name(s):
    return re.sub(r'[\\/*?:"<>|]', "", s).strip()


def parse_month_arg(month_str):
    """Accepts 'April 2027' -> (4, 2027)"""
    dt = datetime.strptime(month_str.strip(), "%B %Y")
    return dt.month, dt.year


def load_tracker_rows():
    wb = openpyxl.load_workbook(TRACKER, data_only=True)
    ws = wb.active
    header_row = 5
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def load_rates():
    wb = openpyxl.load_workbook(MASTER_RATES, data_only=True)
    ws = wb.active
    header_row = 3
    rates_by_level = {}
    for r in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        if not r[0]:
            continue
        price_level, category, item, code, rate, period, source = r
        rates_by_level.setdefault(price_level, []).append({
            "Category": category, "Item": item, "ChargeCode": code,
            "Rate": rate, "Period": period,
        })
    return rates_by_level


def build_rate_schedule_xlsx(price_level, client_name, rate_rows, period_label, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rate Schedule"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Regional Rate Schedule")
    t.font = Font(name=FONT, size=16, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    s = ws.cell(row=2, column=1, value=f"{client_name}   |   Price Level: {price_level}   |   Rates {period_label}")
    s.font = Font(name=FONT, size=11, bold=True, color=WHITE)
    s.fill = PatternFill("solid", fgColor=STEEL)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    header_row = 4
    headers = ["Records Management", "Charge Code", "Rate"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 20

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_num = header_row + 1
    current_category = None
    for item in rate_rows:
        if item["Category"] != current_category:
            current_category = item["Category"]
            cat_cell = ws.cell(row=row_num, column=1, value=current_category)
            cat_cell.font = Font(name=FONT, size=10.5, bold=True, color="C00000")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
            row_num += 1
        ws.cell(row=row_num, column=1, value=item["Item"]).font = Font(name=FONT, size=10)
        code_cell = ws.cell(row=row_num, column=2, value=item["ChargeCode"])
        code_cell.font = Font(name=FONT, size=10)
        code_cell.alignment = Alignment(horizontal="center")
        rate_cell = ws.cell(row=row_num, column=3, value=item["Rate"])
        rate_cell.number_format = '"R" #,##0.00'
        rate_cell.font = Font(name=FONT, size=10, bold=True, color="C00000")
        rate_cell.alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = border
        row_num += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(out_path)


def build_letter_docx(client_name, price_level, effective_date, letter_date, out_path):
    """Clone letter_template.docx and fill in the {{PLACEHOLDER}} fields."""
    tmp_dir = out_path + "__tmp"
    if os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir)
    os.makedirs(tmp_dir)
    with zipfile.ZipFile(LETTER_TEMPLATE) as z:
        z.extractall(tmp_dir)

    doc_xml_path = os.path.join(tmp_dir, "word", "document.xml")
    with open(doc_xml_path, "r", encoding="utf-8") as f:
        xml = f.read()

    replacements = {
        "{{LETTER_DATE}}": letter_date.strftime("%d %B %Y").upper(),
        "{{CLIENT_NAME}}": client_name,
        "{{PRICE_LEVEL}}": price_level,
        "{{EFFECTIVE_DATE_LONG}}": effective_date.strftime("%d %B %Y"),
        "{{EFFECTIVE_DAY}}": str(effective_date.day),
        "{{EFFECTIVE_MONTH}}": effective_date.strftime("%B"),
        "{{EFFECTIVE_YEAR}}": str(effective_date.year),
    }
    for k, v in replacements.items():
        xml = xml.replace(k, v)

    with open(doc_xml_path, "w", encoding="utf-8") as f:
        f.write(xml)

    if os.path.exists(out_path):
        os.remove(out_path)
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(tmp_dir):
            for file in files:
                fp = os.path.join(root, file)
                arcname = os.path.relpath(fp, tmp_dir)
                zf.write(fp, arcname)
    shutil.rmtree(tmp_dir)


def build_manifest(entries, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Manifest"
    ws.sheet_view.showGridLines = False

    headers = ["Sales Person", "Sales Person Email", "Client Name", "Price Level",
               "Effective Date", "Subject", "Body", "Letter Attachment", "Schedule Attachment"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value="EMAIL MANIFEST \u2013 Ready for send_via_outlook.py")
    t.font = Font(name=FONT, size=14, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    header_row = 2
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=STEEL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, e in enumerate(entries):
        row_num = header_row + 1 + i
        band = PatternFill("solid", fgColor=WHITE if i % 2 == 0 else LIGHT_GREY)
        vals = [e["sp_name"], e["sp_email"], e["client"], e["price_level"],
                e["effective_date"].strftime("%d %B %Y"), e["subject"], e["body"],
                e["letter_path"], e["schedule_path"]]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.font = Font(name=FONT, size=9.5)
            cell.fill = band
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))

    widths = {1: 16, 2: 26, 3: 30, 4: 13, 5: 15, 6: 30, 7: 45, 8: 40, 9: 40}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(out_path)


def generate_pack(month_str):
    """Generate the letter+schedule pack for the given month (e.g. 'April 2027').
    Returns the list of manifest entry dicts. Writes files under OUTPUT_DIR."""
    target_month, target_year = parse_month_arg(month_str)

    rows = load_tracker_rows()
    rates_by_level = load_rates()

    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    os.makedirs(OUTPUT_DIR)

    matched = []
    for r in rows:
        increase_month = r["Rate Increase Month"]
        if not isinstance(increase_month, (datetime, date)):
            continue
        due_month = increase_month.month - 3
        due_year = increase_month.year
        if due_month <= 0:
            due_month += 12
            due_year -= 1
        if (due_month, due_year) != (target_month, target_year):
            continue
        matched.append(r)

    manifest_entries = []
    letter_date = date.today()

    for r in matched:
        price_level = r["Price Level"]
        client_name = r["Client Name"]
        sp_name = r["Sales Person"]
        sp_email = r["Sales Person Email"]
        effective_date = r["Rate Increase Month"]
        if isinstance(effective_date, datetime):
            effective_date = effective_date.date()

        rate_rows = rates_by_level.get(price_level) or rates_by_level.get(STANDARD_LEVEL)
        is_placeholder = price_level not in rates_by_level or price_level == STANDARD_LEVEL
        period_label = rate_rows[0]["Period"] if rate_rows else ""

        client_folder = os.path.join(OUTPUT_DIR, safe_name(sp_name), safe_name(client_name))
        os.makedirs(client_folder, exist_ok=True)

        letter_path = os.path.join(client_folder, f"Price Increase Letter - {safe_name(client_name)}.docx")
        schedule_path = os.path.join(client_folder, f"Rate Schedule - {safe_name(client_name)}.xlsx")

        build_letter_docx(client_name, price_level, effective_date, letter_date, letter_path)
        build_rate_schedule_xlsx(price_level, client_name, rate_rows, period_label, schedule_path)

        subject = f"Action required: Rate increase notification due - {client_name} ({price_level})"
        body = (f"Hi {sp_name.split()[0]},\n\n"
                f"{client_name}'s rate increase takes effect {effective_date.strftime('%d %B %Y')}. "
                f"Attached are the price increase letter and rate schedule for you to forward to the client "
                f"ahead of the notice deadline.\n\nThanks,\nPricing Team")

        manifest_entries.append({
            "sp_name": sp_name, "sp_email": sp_email, "client": client_name,
            "price_level": price_level, "effective_date": effective_date,
            "subject": subject, "body": body,
            "letter_path": os.path.abspath(letter_path),
            "schedule_path": os.path.abspath(schedule_path),
            "rate_rows": rate_rows,
            "placeholder_rates": bool(rate_rows) and rate_rows[0].get("Source", "").startswith("DEMO"),
        })

    manifest_path = os.path.join(OUTPUT_DIR, "email_manifest.xlsx")
    build_manifest(manifest_entries, manifest_path)
    return manifest_entries


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", required=True, help="Target notification month, e.g. 'April 2027'")
    args = parser.parse_args()
    entries = generate_pack(args.month)
    print(f"Generated {len(entries)} letter+schedule pairs -> {OUTPUT_DIR}/")
    print(f"Manifest written -> {OUTPUT_DIR}/email_manifest.xlsx")


if __name__ == "__main__":
    main()
